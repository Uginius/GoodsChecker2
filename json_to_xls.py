import json
import math
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font

from parsers.parse_rating_votonia import VotoniaRatingUpdate
from utilites import check_dir


class XlsConverter:
    def __init__(self, filename):
        self.load_filename = f'json_results/{filename}'
        self.goods_dict = {}
        self.workbook = Workbook()
        self.date = filename.split('.')[0]

    def run(self):
        self.open_json()
        self.initiate_workbook()
        self.combine_data_to_write()
        check_dir('xls_results')
        self.workbook.save(f"xls_results/{self.date}.xlsx")

    def open_json(self):
        with open(self.load_filename, 'r', encoding='utf8') as json_file:
            self.goods_dict = json.load(json_file)

    def combine_data_to_write(self):
        sw = self.workbook.active
        for shop in self.goods_dict:
            self.check_empty_platform_ratings(shop)
            shop_goods = self.goods_dict[shop]
            for shop_id in shop_goods:
                merch = shop_goods[shop_id]
                qt = merch['vote_qt']
                rating = merch['vote_rating']
                if rating is None:
                    need = None
                else:
                    if rating > 4 and qt > 1:
                        need = 0
                    else:
                        need = math.ceil(qt * abs(4 - rating)) if rating else 2
                data = [shop, int(shop_id), merch['brand'], merch['name'], merch['url'], merch['status'],
                        merch['price'], rating, qt, need]
                sw.append(data)
                self.conditions_and_cell_styles()

    def initiate_workbook(self):
        title = ['shop', 'shop id', 'brand', 'name', 'url', 'status',
                 'price', 'rating', 'votes', 'need']
        ws = self.workbook.active
        ws.append(title)
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 15
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor='b3b3b3')
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def conditions_and_cell_styles(self):
        sw = self.workbook.active
        cur_row = sw.max_row
        url_cell = sw[f'E{cur_row}']
        url_cell.font = Font(size=7)
        rating_cell = sw[f'H{cur_row}']
        votes_cell = sw[f'I{cur_row}']
        need_votes_cell = sw[f'J{cur_row}']
        rating_cell.alignment = Alignment(horizontal='center', vertical='center')
        votes_cell.alignment = Alignment(horizontal='center', vertical='center')
        need_votes_cell.alignment = Alignment(horizontal='center', vertical='center')
        if rating_cell.value:
            if rating_cell.value == 0:
                rating_cell.fill = PatternFill("solid", fgColor='ffe4e1')
            elif rating_cell.value < 4:
                rating_cell.fill = PatternFill("solid", fgColor='E6B8B7')
        if need_votes_cell.value:
            if need_votes_cell.value > 0:
                need_votes_cell.fill = PatternFill("solid", fgColor='ffcd75')

    def check_empty_platform_ratings(self, shop):
        match shop:
            case 'votonia':
                votonia_goods_updater = VotoniaRatingUpdate(self.goods_dict[shop])
                votonia_goods_updater.run()
                self.goods_dict[shop] = votonia_goods_updater.out_goods_data()
            case 'maxidom':
                pass
